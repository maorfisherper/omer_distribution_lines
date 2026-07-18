import csv
import math
import random
import re
from itertools import permutations

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

SRC_CSV = "branches_with_waze_links_v3.csv"
OUT_XLSX = "branches_with_waze_links_v3_days.xlsx"
MIN_PER_DAY = 7
MAX_PER_DAY = 9
R_EARTH = 6371000.0


def haversine_m(lat1, lon1, lat2, lon2):
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dlambda / 2) ** 2
    return 2 * R_EARTH * math.asin(math.sqrt(a))


rows = list(csv.reader(open(SRC_CSV, encoding="utf-8-sig")))
header, data = rows[0], rows[1:]

coord_re = re.compile(r"^\s*(-?\d+(?:\.\d+)?)\s*,\s*(-?\d+(?:\.\d+)?)\s*$")

valid = []
invalid = []
for row in data:
    m = coord_re.match(row[3]) if row[3] else None
    if m:
        lat, lon = float(m.group(1)), float(m.group(2))
        valid.append({"row": row, "lat": lat, "lon": lon})
    else:
        invalid.append(row)

n = len(valid)
print("valid points:", n, "invalid/no-coord:", len(invalid))

mean_lat = sum(p["lat"] for p in valid) / n
mean_lat_rad = math.radians(mean_lat)
for p in valid:
    p["x"] = math.radians(p["lon"]) * math.cos(mean_lat_rad) * R_EARTH
    p["y"] = math.radians(p["lat"]) * R_EARTH

D = [[0.0] * n for _ in range(n)]
for i in range(n):
    for j in range(i + 1, n):
        d = haversine_m(valid[i]["lat"], valid[i]["lon"], valid[j]["lat"], valid[j]["lon"])
        D[i][j] = D[j][i] = d


def planar_dist2(a, b):
    dx, dy = a["x"] - b["x"], a["y"] - b["y"]
    return dx * dx + dy * dy


def recompute_center(members_idx):
    x = sum(valid[i]["x"] for i in members_idx) / len(members_idx)
    y = sum(valid[i]["y"] for i in members_idx) / len(members_idx)
    return {"x": x, "y": y}


# ---- feasible day counts given 7-9 points/day and n points ----
feasible_k = [k for k in range(math.ceil(n / MAX_PER_DAY), math.floor(n / MIN_PER_DAY) + 1)]
print("feasible day counts:", feasible_k)


def constrained_kmeans(k, seed):
    rng = random.Random(seed)
    centers = [dict(x=valid[rng.randrange(n)]["x"], y=valid[rng.randrange(n)]["y"])]
    while len(centers) < k:
        dists = [min(planar_dist2(p, c) for c in centers) for p in valid]
        total = sum(dists)
        r = rng.uniform(0, total) if total > 0 else 0
        acc, chosen = 0.0, None
        for i, d in enumerate(dists):
            acc += d
            if acc >= r:
                chosen = i
                break
        if chosen is None:
            chosen = rng.randrange(n)
        centers.append({"x": valid[chosen]["x"], "y": valid[chosen]["y"]})

    assign = [0] * n
    for _ in range(50):
        changed = False
        for i, p in enumerate(valid):
            best_c = min(range(k), key=lambda ci: planar_dist2(p, centers[ci]))
            if assign[i] != best_c:
                assign[i] = best_c
                changed = True
        for ci in range(k):
            members = [i for i in range(n) if assign[i] == ci]
            if members:
                centers[ci] = recompute_center(members)
        if not changed:
            break

    return {ci: [i for i in range(n) if assign[i] == ci] for ci in range(k)}


def enforce_capacity(clusters, min_size=MIN_PER_DAY, max_size=MAX_PER_DAY, max_iter=2000):
    clusters = {ci: list(v) for ci, v in clusters.items()}

    def center_of(ci):
        return recompute_center(clusters[ci]) if clusters[ci] else None

    for _ in range(max_iter):
        over = [ci for ci in clusters if len(clusters[ci]) > max_size]
        if over:
            ci = over[0]
            c = center_of(ci)
            worst = max(clusters[ci], key=lambda i: planar_dist2(valid[i], c))
            candidates = [cj for cj in clusters if cj != ci and len(clusters[cj]) < max_size]
            if not candidates:
                break

            def key_fn(cj):
                cc = center_of(cj)
                return 0.0 if cc is None else planar_dist2(valid[worst], cc)

            target = min(candidates, key=key_fn)
            clusters[ci].remove(worst)
            clusters[target].append(worst)
            continue

        under = [ci for ci in clusters if len(clusters[ci]) < min_size]
        if under:
            ci = under[0]
            c = center_of(ci)
            donors = [
                (cj, i) for cj in clusters if cj != ci and len(clusters[cj]) > min_size for i in clusters[cj]
            ]
            if not donors:
                break
            if c is None:
                cj, i = donors[0]
            else:
                cj, i = min(donors, key=lambda t: planar_dist2(valid[t[1]], c))
            clusters[cj].remove(i)
            clusters[ci].append(i)
            continue
        break

    return clusters


def swap_refine(clusters, passes=10):
    clusters = {ci: list(v) for ci, v in clusters.items()}
    ids = list(clusters.keys())
    for _ in range(passes):
        improved = False
        centers = {ci: recompute_center(clusters[ci]) for ci in ids}
        for a_idx in range(len(ids)):
            for b_idx in range(a_idx + 1, len(ids)):
                ca, cb = ids[a_idx], ids[b_idx]
                for i in list(clusters[ca]):
                    for j in list(clusters[cb]):
                        cur = planar_dist2(valid[i], centers[ca]) + planar_dist2(valid[j], centers[cb])
                        new = planar_dist2(valid[j], centers[ca]) + planar_dist2(valid[i], centers[cb])
                        if new < cur - 1e-6:
                            clusters[ca].remove(i)
                            clusters[ca].append(j)
                            clusters[cb].remove(j)
                            clusters[cb].append(i)
                            centers[ca] = recompute_center(clusters[ca])
                            centers[cb] = recompute_center(clusters[cb])
                            improved = True
                            break
                    else:
                        continue
                    break
        if not improved:
            break
    return clusters


def optimal_path_order(indices):
    if len(indices) <= 1:
        return list(indices), 0.0
    best_order, best_len = None, None
    for perm in permutations(indices):
        length = sum(D[perm[i]][perm[i + 1]] for i in range(len(perm) - 1))
        if best_len is None or length < best_len:
            best_len, best_order = length, list(perm)
    return best_order, best_len


def total_route_length(clusters):
    total = 0.0
    orders = {}
    for ci, members in clusters.items():
        order, length = optimal_path_order(members)
        orders[ci] = (order, length)
        total += length
    return total, orders


RESTARTS = 12
results_by_k = {}
for k in feasible_k:
    best_clusters, best_total, best_orders = None, None, None
    for seed in range(RESTARTS):
        clusters = constrained_kmeans(k, seed)
        clusters = enforce_capacity(clusters)
        if not all(MIN_PER_DAY <= len(v) <= MAX_PER_DAY for v in clusters.values()):
            continue
        clusters = swap_refine(clusters)
        if not all(MIN_PER_DAY <= len(v) <= MAX_PER_DAY for v in clusters.values()):
            continue
        total, orders = total_route_length(clusters)
        if best_total is None or total < best_total:
            best_total, best_clusters, best_orders = total, clusters, orders
    results_by_k[k] = (best_total, best_clusters, best_orders)
    print(f"k={k}: best total distance = {best_total:.0f} m ({best_total/1000:.1f} km), sizes={sorted(len(v) for v in best_clusters.values())}")

best_k = min(results_by_k, key=lambda k: results_by_k[k][0])
best_total, best_clusters, best_orders = results_by_k[best_k]
print(f"\nchosen: {best_k} days, total {best_total/1000:.1f} km")

day_results = []


def centroid(members_idx):
    c = recompute_center(members_idx)
    return c["x"], c["y"]


ordered_cluster_ids = sorted(best_clusters.keys(), key=lambda ci: -centroid(best_clusters[ci])[1])
for day_num, ci in enumerate(ordered_cluster_ids, start=1):
    order, length = best_orders[ci]
    day_results.append((day_num, order, length))
    print(f"day {day_num}: {len(order)} stops, {length:.0f} m total")

# ================= build output workbook =================
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Route"
ws.sheet_view.rightToLeft = True
ws.freeze_panes = "A2"

out_header = ["Day", "Stop", "Original text", "Corrected text", "Waze location",
              "Coordinates", "Leg distance from previous stop (m)"]
ws.append(out_header)

header_font = Font(b=True)
header_fill = PatternFill(patternType="solid", fgColor="00D9E1F2")
for c in range(1, len(out_header) + 1):
    cell = ws.cell(1, c)
    cell.font = header_font
    cell.fill = header_fill

link_font = Font(color="000563C1", u="single")
wrap_center = Alignment(vertical="center", wrapText=True)

excel_row = 2
for day_num, order, length in day_results:
    for stop_num, pt_idx in enumerate(order, start=1):
        row = valid[pt_idx]["row"]
        leg = None
        if stop_num > 1:
            prev_idx = order[stop_num - 2]
            leg = round(D[prev_idx][pt_idx])
        ws.cell(excel_row, 1, day_num)
        ws.cell(excel_row, 2, stop_num)
        ws.cell(excel_row, 3, row[0])
        ws.cell(excel_row, 4, row[1])
        wcell = ws.cell(excel_row, 5, row[2])
        if isinstance(row[2], str) and row[2].startswith("http"):
            wcell.hyperlink = row[2]
            wcell.font = link_font
        ws.cell(excel_row, 6, row[3])
        ws.cell(excel_row, 7, leg if leg is not None else None)
        for c in range(1, 8):
            ws.cell(excel_row, c).alignment = wrap_center
        excel_row += 1

for row in invalid:
    ws.cell(excel_row, 1, "Unscheduled")
    ws.cell(excel_row, 2, None)
    ws.cell(excel_row, 3, row[0])
    ws.cell(excel_row, 4, row[1])
    ws.cell(excel_row, 5, row[2])
    ws.cell(excel_row, 6, row[3])
    ws.cell(excel_row, 7, None)
    for c in range(1, 8):
        ws.cell(excel_row, c).alignment = wrap_center
    excel_row += 1

widths = {"A": 12, "B": 8, "C": 30, "D": 42, "E": 46, "F": 22, "G": 30}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

ws2 = wb.create_sheet("Summary")
ws2.sheet_view.rightToLeft = True
ws2.append(["Day", "Stops", "Total distance (m)", "Total distance (km)"])
for c in range(1, 5):
    cell = ws2.cell(1, c)
    cell.font = header_font
    cell.fill = header_fill
for day_num, order, length in day_results:
    ws2.append([day_num, len(order), round(length), round(length / 1000, 1)])
ws2.append(["Unscheduled", len(invalid), None, None])
ws2.append([])
ws2.append(["Day-count option", "Total distance (km)"])
for k in sorted(results_by_k):
    t = results_by_k[k][0]
    ws2.append([k, round(t / 1000, 1) if t is not None else None])
for col, w in {"A": 18, "B": 10, "C": 18, "D": 18}.items():
    ws2.column_dimensions[col].width = w

wb.save(OUT_XLSX)
print("wrote", OUT_XLSX)
