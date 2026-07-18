import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

SRC_CSV = "branches_with_waze_links_v3.csv"
OUT_XLSX = "branches_with_waze_links_v3.xlsx"

rows = list(csv.reader(open(SRC_CSV, encoding="utf-8-sig")))
header, data = rows[0], rows[1:]

wb = openpyxl.Workbook()
ws = wb.active
ws.sheet_view.rightToLeft = True
ws.freeze_panes = "A2"
ws.append(header)

header_font = Font(b=True)
header_fill = PatternFill(patternType="solid", fgColor="00D9E1F2")
for c in range(1, len(header) + 1):
    cell = ws.cell(1, c)
    cell.font = header_font
    cell.fill = header_fill

link_font = Font(color="000563C1", u="single")
wrap_center = Alignment(vertical="center", wrapText=True)

for r, row in enumerate(data):
    excel_row = r + 2
    for c, val in enumerate(row):
        cell = ws.cell(excel_row, c + 1)
        if c in (4, 5) and val != "":
            cell.value = int(val)
        else:
            cell.value = val if val != "" else None
        cell.alignment = wrap_center
        if c == 2 and isinstance(val, str) and val.startswith("http"):
            cell.hyperlink = val
            cell.font = link_font

widths = {"A": 30.0, "B": 42.0, "C": 46.0, "D": 22.0, "E": 16.0, "F": 16.0}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

wb.save(OUT_XLSX)
print("wrote", OUT_XLSX)
