import math
import re
from numbers_parser import Document
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SRC_NUMBERS = "branches_with_waze_links_v2.numbers"
SRC_REVIEW = "branches_with_waze_links_v2_REVIEW.xlsx"
OUT_NUMBERS = "branches_with_waze_links_v3.numbers"
OUT_CSV = "branches_with_waze_links_v3.csv"
OUT_REVIEW = "branches_with_waze_links_v3_REVIEW.xlsx"

def haversine_m(lat1, lon1, lat2, lon2):
    R = 6371000.0
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dlambda / 2) ** 2
    return 2 * R * math.asin(math.sqrt(a))

# --- read source data from the .numbers deliverable ---
doc = Document(SRC_NUMBERS)
table = doc.sheets[0].tables[0]
headers = [table.cell(0, c).value for c in range(table.num_cols)]
rows = []
for r in range(1, table.num_rows):
    rows.append([table.cell(r, c).value for c in range(table.num_cols)])

# --- read Note column from the REVIEW xlsx (same row order) ---
wb_src = openpyxl.load_workbook(SRC_REVIEW)
ws_src = wb_src.active
notes = []
for r in range(2, ws_src.max_row + 1):
    notes.append(ws_src.cell(r, 5).value)

assert len(notes) == len(rows), (len(notes), len(rows))

coord_re = re.compile(r"^\s*(-?\d+(?:\.\d+)?)\s*,\s*(-?\d+(?:\.\d+)?)\s*$")

coords = []  # (lat, lon) or None per row
for row in rows:
    raw = row[3]
    m = coord_re.match(raw) if isinstance(raw, str) else None
    if m:
        coords.append((float(m.group(1)), float(m.group(2))))
    else:
        coords.append(None)

n = len(rows)
closest_index = [None] * n   # sheet row number (header=1, first data row=2)
closest_meters = [None] * n

for i in range(n):
    if coords[i] is None:
        continue
    best_j, best_d = None, None
    for j in range(n):
        if j == i or coords[j] is None:
            continue
        d = haversine_m(coords[i][0], coords[i][1], coords[j][0], coords[j][1])
        if best_d is None or d < best_d:
            best_d, best_j = d, j
    if best_j is not None:
        closest_index[i] = best_j + 2  # +1 for header row, +1 for 1-based
        closest_meters[i] = round(best_d)

new_headers = headers + ["closest_arial_distance_index", "closest_arial_distance_meters"]

# ================= .numbers =================
table.add_column(num_cols=2)
idx_col = table.num_cols - 2
m_col = table.num_cols - 1
table.write(0, idx_col, new_headers[-2])
table.write(0, m_col, new_headers[-1])
for r in range(n):
    idx_val = closest_index[r]
    m_val = closest_meters[r]
    table.write(r + 1, idx_col, idx_val if idx_val is not None else "")
    table.write(r + 1, m_col, m_val if m_val is not None else "")
    if idx_val is not None:
        table.set_cell_formatting(r + 1, idx_col, "number", decimal_places=0)
        table.set_cell_formatting(r + 1, m_col, "number", decimal_places=0)
doc.save(OUT_NUMBERS)
print("wrote", OUT_NUMBERS)

# ================= .csv =================
import csv
with open(OUT_CSV, "w", newline="", encoding="utf-8-sig") as f:
    w = csv.writer(f)
    w.writerow(new_headers)
    for r in range(n):
        row = list(rows[r])
        row.append(closest_index[r] if closest_index[r] is not None else "")
        row.append(closest_meters[r] if closest_meters[r] is not None else "")
        w.writerow(row)
print("wrote", OUT_CSV)

# ================= REVIEW.xlsx =================
wb = openpyxl.Workbook()
ws = wb.active
ws.sheet_view.rightToLeft = True
ws.freeze_panes = "A2"

full_headers = headers + ["closest_arial_distance_index", "closest_arial_distance_meters", "Note"]
ws.append(full_headers)

header_font = Font(b=True)
header_fill = PatternFill(patternType="solid", fgColor="00D9E1F2")
for c in range(1, len(full_headers) + 1):
    cell = ws.cell(1, c)
    cell.font = header_font
    cell.fill = header_fill

link_font = Font(color="000563C1", u="single")
wrap_center = Alignment(vertical="center", wrapText=True)

for r in range(n):
    excel_row = r + 2
    row = rows[r]
    for c in range(4):  # Original text, Corrected text, Waze location, Coordinates
        cell = ws.cell(excel_row, c + 1)
        cell.value = row[c]
        cell.alignment = wrap_center
        if c == 2 and isinstance(row[c], str) and row[c].startswith("http"):
            cell.hyperlink = row[c]
            cell.font = link_font
    idx_val = closest_index[r]
    m_val = closest_meters[r]
    ci = ws.cell(excel_row, 5)
    ci.value = idx_val if idx_val is not None else None
    ci.alignment = wrap_center
    cm = ws.cell(excel_row, 6)
    cm.value = m_val if m_val is not None else None
    cm.alignment = wrap_center
    cn = ws.cell(excel_row, 7)
    cn.value = notes[r]
    cn.alignment = wrap_center

widths = {"A": 30.0, "B": 42.0, "C": 46.0, "D": 22.0, "E": 16.0, "F": 16.0, "G": 56.0}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

wb.save(OUT_REVIEW)
print("wrote", OUT_REVIEW)

# quick sanity print
for r in range(n):
    if coords[r] is None:
        print("NO COORD:", rows[r][1])
