import json
import os
import re
import urllib.parse

import openpyxl

DAYS_XLSX = "branches_with_waze_links_v3_days.xlsx"
ALL_XLSX = "branches_with_waze_links_v3.xlsx"
DRIVING_JSON = "driving_routes.json"
MATRIX_JSON = "driving_matrix.json"
OUT_JS = "docs/data.js"

driving = {}
if os.path.exists(DRIVING_JSON):
    with open(DRIVING_JSON, encoding="utf-8") as f:
        driving = {int(k): v for k, v in json.load(f).items()}
else:
    print(f"warning: {DRIVING_JSON} not found — run scripts/fetch_driving_distances.py first; "
          "site will fall back to straight-line distances only")

with open(MATRIX_JSON, encoding="utf-8") as f:
    matrix_data = json.load(f)
DEPOT = matrix_data["depot"]


def google_maps_url(real_stops):
    depot_pt = f"{DEPOT['lat']},{DEPOT['lon']}"
    pts = [f"{s['coordinates']['lat']},{s['coordinates']['lon']}" for s in real_stops if s["coordinates"]]
    if not pts:
        return None
    params = {"api": "1", "origin": depot_pt, "destination": depot_pt, "travelmode": "driving"}
    params["waypoints"] = "|".join(pts)
    return "https://www.google.com/maps/dir/?" + urllib.parse.urlencode(params, safe="|,")


coord_re = re.compile(r"^\s*(-?\d+(?:\.\d+)?)\s*,\s*(-?\d+(?:\.\d+)?)\s*$")


def parse_coords(s):
    if not s:
        return None
    m = coord_re.match(s)
    if not m:
        return None
    return {"lat": float(m.group(1)), "lon": float(m.group(2))}


DEPOT_COORDS = {"lat": DEPOT["lat"], "lon": DEPOT["lon"]}

# ---- day routes ----
wb = openpyxl.load_workbook(DAYS_XLSX, data_only=True)
route_ws = wb["Route"]

days = {}  # day_num -> list of real-branch stop entries
depot_return_leg = {}  # day_num -> leg distance (m) from the synthetic "return to depot" row
unscheduled = []
for row in route_ws.iter_rows(min_row=2, values_only=True):
    day, stop, orig, corrected, waze, coords, leg = row
    if day == "Unscheduled":
        unscheduled.append({
            "originalText": orig,
            "correctedText": corrected,
            "wazeUrl": waze,
            "coordinates": parse_coords(coords),
        })
        continue
    if not orig:  # synthetic "return to depot" bookend row
        depot_return_leg[int(day)] = int(leg) if leg is not None else None
        continue
    days.setdefault(int(day), []).append({
        "originalText": orig,
        "correctedText": corrected,
        "wazeUrl": waze,
        "coordinates": parse_coords(coords),
        "legDistanceM": int(leg) if leg is not None else None,
        "stop": int(stop),
    })

day_list = []
for day_num in sorted(days):
    real_stops = sorted(days[day_num], key=lambda s: s["stop"])
    drv = driving.get(day_num)

    if drv:
        for i, s in enumerate(real_stops):
            leg = drv["legs"][i]  # legs[0] = depot->stop1, legs[i] = stop_i -> stop_(i+1)
            s["legDrivingDistanceM"] = leg["distanceM"]
            s["legDrivingDurationS"] = leg["durationS"]
        return_leg = drv["legs"][-1]
        return_leg_m, return_leg_s = return_leg["distanceM"], return_leg["durationS"]
    else:
        return_leg_m, return_leg_s = depot_return_leg.get(day_num), None

    depot_start = {
        "isDepot": True, "role": "start",
        "originalText": None, "correctedText": DEPOT["name"],
        "wazeUrl": DEPOT["wazeUrl"], "coordinates": DEPOT_COORDS,
        "legDistanceM": None, "legDrivingDistanceM": None, "legDrivingDurationS": None,
    }
    depot_end = {
        "isDepot": True, "role": "end",
        "originalText": None, "correctedText": DEPOT["name"] + " (חזרה)",
        "wazeUrl": DEPOT["wazeUrl"], "coordinates": DEPOT_COORDS,
        "legDistanceM": depot_return_leg.get(day_num),
        "legDrivingDistanceM": return_leg_m, "legDrivingDurationS": return_leg_s,
    }
    stops = [depot_start] + real_stops + [depot_end]

    total = sum(s["legDistanceM"] for s in stops if s.get("legDistanceM"))

    day_entry = {
        "day": day_num,
        "stops": stops,
        "totalDistanceM": total,
        "drivingDistanceM": drv["totalDistanceM"] if drv else None,
        "drivingDurationS": drv["totalDurationS"] if drv else None,
        "routeGeometry": drv["geometry"] if drv else None,
        "googleMapsUrl": google_maps_url(real_stops),
    }
    day_list.append(day_entry)

# ---- full branch database ----
wb2 = openpyxl.load_workbook(ALL_XLSX, data_only=True)
all_ws = wb2.active

# map (originalText) -> (day, stop) for cross-referencing (real branch stops only, not depot bookends)
schedule_lookup = {}
for d in day_list:
    for s in d["stops"]:
        if not s.get("isDepot"):
            schedule_lookup[s["originalText"]] = {"day": d["day"], "stop": s["stop"]}

all_branches = []
for row in all_ws.iter_rows(min_row=2, values_only=True):
    orig, corrected, waze, coords, closest_idx, closest_m = row
    sched = schedule_lookup.get(orig)
    all_branches.append({
        "originalText": orig,
        "correctedText": corrected,
        "wazeUrl": waze,
        "coordinates": parse_coords(coords),
        "closestDistanceM": int(closest_m) if closest_m is not None else None,
        "day": sched["day"] if sched else None,
        "stop": sched["stop"] if sched else None,
    })

# ---- planner data: full driving-distance matrix + depot, for client-side route recomputation ----
branch_lookup = {b["originalText"]: b for b in all_branches}
planner_points = []
for ot in matrix_data["originalTexts"]:
    b = branch_lookup[ot]
    planner_points.append({
        "originalText": b["originalText"],
        "correctedText": b["correctedText"],
        "wazeUrl": b["wazeUrl"],
        "coordinates": b["coordinates"],
    })

data = {
    "generated": True,
    "depot": {"name": DEPOT["name"], "coordinates": DEPOT_COORDS, "wazeUrl": DEPOT["wazeUrl"]},
    "days": day_list,
    "unscheduled": unscheduled,
    "allBranches": all_branches,
    "planner": {
        "points": planner_points,
        "distances": matrix_data["distances"],  # meters, index 0 = depot, 1..N = planner_points order
        "durations": matrix_data["durations"],  # seconds, same indexing
    },
}

with open(OUT_JS, "w", encoding="utf-8") as f:
    f.write("// Auto-generated by scripts/build_site_data.py — do not edit by hand.\n")
    f.write("window.APP_DATA = ")
    f.write(json.dumps(data, ensure_ascii=False, indent=2))
    f.write(";\n")

real_stop_count = sum(len(d["stops"]) - 2 for d in day_list)
print(f"wrote {OUT_JS}: {len(day_list)} days, {real_stop_count} scheduled stops, "
      f"{len(unscheduled)} unscheduled, {len(all_branches)} total branches")
